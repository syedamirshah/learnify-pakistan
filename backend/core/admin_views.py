import openpyxl
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .forms import UploadForm
from .models import User
from django.utils import timezone
from datetime import timedelta
from django.db.models import Case, When, IntegerField
from django.utils.html import format_html
from .forms import UploadSCQForm
from .models import SCQQuestion, QuestionBank
from core.forms import UploadSCQForm, UploadMCQForm, UploadFIBForm
from core.models import SCQQuestion, MCQQuestion, FIBQuestion, QuestionBank , QuizQuestionAssignment , Quiz
from django.contrib.admin.views.decorators import staff_member_required
import uuid
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseForbidden
from .models import Quiz, QuizQuestionAssignment
import os
from django.contrib.auth.decorators import user_passes_test
from django.conf import settings
from django.http import FileResponse, Http404
from .models import Grade, Subject, Chapter
from django.db.models.functions import Lower
from django import forms
from django.core.paginator import Paginator
from core.utils import send_account_notification_email  # ‚úÖ Add this at the top
from django.db.models import Count, OuterRef, Subquery, IntegerField, Value, Case, When
from django.views.decorators.http import require_POST
from django.contrib.auth import get_user_model
from django.http import JsonResponse
from django.http import FileResponse, HttpResponseRedirect
from django.core.management import call_command
from datetime import date  # ‚úÖ add this import at the top






def bulk_upload_students(request):
    if request.method == 'POST':
        form = UploadForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['excel_file']
            wb = openpyxl.load_workbook(file)
            sheet = wb.active

            uploaded_count = 0
            skipped_count = 0

            for row in sheet.iter_rows(min_row=2, values_only=True):
                try:
                    (
                        username, full_name, language_used_at_home, email, password, role, gender,
                        schooling_status, grade_name, school_name, city, province,
                        subscription_plan
                    ) = row[:13]

                    if User.objects.filter(username=username).exists():
                        skipped_count += 1
                        continue

                    # ✅ Convert grade name to Grade instance
                    try:
                        cleaned_grade_name = str(grade_name).strip().replace('"', '').replace("'", '')
                        grade_instance = Grade.objects.get(name=grade_name)
                    except Grade.DoesNotExist:
                        messages.error(request, f"⛔ Error in row {row}: Grade '{grade_name}' does not exist in database.")
                        continue

                    user = User.objects.create(
                        username=username,
                        full_name=full_name,
                        email=email,
                        role=role,
                        gender=gender,
                        schooling_status=schooling_status,
                        grade=grade_instance,
                        school_name=school_name,
                        city=city,
                        province=province,
                        subscription_plan=subscription_plan,
                        language_used_at_home=language_used_at_home or '',
                        account_status='inactive'
                    )
                    user.set_password(password)  # ✅ Hash password
                    user.save()

                    uploaded_count += 1

                except Exception as e:
                    messages.error(request, f"⚠️ Error in row {row}: {e}")

            messages.success(
                request,
                f"✅ Uploaded {uploaded_count} students. Skipped {skipped_count} existing usernames."
            )
            return redirect('/admin/core/user/complete_user_data/')
    else:
        form = UploadForm()

    return render(request, 'admin/core/student_upload_form.html', {'form': form})


def manage_subscriptions(request):
    # ‚úÖ Handle POST actions first
    if request.method == 'POST':
        action = request.POST.get('action')
        user_id = request.POST.get('user_id')
        duration = request.POST.get('duration')

        try:
            user = User.objects.get(id=user_id)

            if action == 'disable':
                user.account_status = 'inactive'
                user.subscription_expiry = None
                user.is_active = False
                user.save()
                messages.success(request, f"User {user.username} disabled successfully.")

            else:
                months = 1 if duration == 'month' else 12
                new_expiry = timezone.now().date() + timedelta(days=30 * months)

                if action == 'activate':
                    user.account_status = 'active'
                    user.subscription_expiry = new_expiry
                    user.is_active = True
                    user.save()

                    # ‚úÖ Send Activation Email
                    send_account_notification_email(user, action='activated')

                elif action == 'extend':
                    if user.subscription_expiry and user.subscription_expiry > timezone.now().date():
                        user.subscription_expiry += timedelta(days=30 * months)
                    else:
                        user.subscription_expiry = new_expiry
                    user.account_status = 'active'
                    user.is_active = True
                    user.save()

                    # ‚úÖ Send Extension Email
                    send_account_notification_email(user, action='extended')

                messages.success(request, f"{action.title()} successful for {user.username} ({duration})")

        except User.DoesNotExist:
            messages.error(request, "User not found.")

        return redirect('manage_subscriptions')

    # Step 1: Filter Users
    users = User.objects.filter(
        role__in=['student', 'teacher', 'manager']
    ).annotate(
        status_order=Case(
            When(account_status='inactive', then=0),
            When(account_status='expired', then=1),
            When(account_status='active', then=2),
            default=3,
            output_field=IntegerField()
        )
    ).order_by('status_order', '-date_joined')

    # Step 2: Handle Filters
    renewal = request.GET.get('renewal')
    status = request.GET.get('status')
    activation = request.GET.get('activation')

    if renewal == 'yes':
        users = users.filter(renewal_requested=True)
    elif renewal == 'no':
        users = users.filter(renewal_requested=False)

    if status:
        users = users.filter(account_status=status)

    if activation == 'yes':
        users = users.filter(account_status='inactive')

    # Step 3: Handle Pagination
    paginate_by = request.GET.get('paginate_by', '50')
    try:
        per_page = len(users) if paginate_by == 'all' else int(paginate_by)
    except ValueError:
        per_page = 50

    paginator = Paginator(users, per_page)
    page_number = request.GET.get('page')
    paginated_users = paginator.get_page(page_number)

    # Step 4: Render Template
    return render(request, 'admin/core/manage_subscriptions.html', {
        'users': paginated_users,
        'renewal': renewal,
        'status': status,
        'activation': activation,
        'paginate_by': paginate_by,
        'request': request  # needed for filter state
    })


def complete_user_data_view(request):
    sort_by = request.GET.get('sort', 'date_joined')
    role_filter = request.GET.get('role')
    province_filter = request.GET.get('province')
    schooling_status_filter = request.GET.get('schooling_status')
    per_page = request.GET.get('per_page', '20')
    page_number = request.GET.get('page')

    users = User.objects.all()

    if role_filter and role_filter != 'All':
        users = users.filter(role=role_filter)
    if province_filter and province_filter != 'All':
        users = users.filter(province=province_filter)
    if schooling_status_filter and schooling_status_filter != 'All':
        users = users.filter(schooling_status=schooling_status_filter)

    users = users.order_by(sort_by)

    # Distinct cleaned-up filter values (excluding duplicate 'All') and ensuring all key roles appear
    expected_roles = {'admin', 'teacher', 'student', 'manager'}

    roles_queryset = User.objects.exclude(role__isnull=True).exclude(role__exact="").values_list('role', flat=True)
    roles_set = set(r.lower() for r in roles_queryset if r.lower() != "all")

    # Union with expected roles to ensure they always appear
    all_roles = roles_set.union(expected_roles)

    # Capitalize roles for consistent dropdown display (optional)
    roles = ['All'] + sorted(all_roles)

    provinces = sorted(set(
        User.objects.exclude(province__isnull=True)
                    .exclude(province__exact="")
                    .exclude(province__iexact="All")
                    .values_list('province', flat=True)
    ))
    provinces.insert(0, 'All')

    schooling_statuses = sorted(set(
        User.objects.exclude(schooling_status__isnull=True)
                    .exclude(schooling_status__exact="")
                    .exclude(schooling_status__iexact="All")
                    .values_list('schooling_status', flat=True)
    ))
    schooling_statuses.insert(0, 'All')

    # Pagination
    if per_page == 'all':
        paginated_users = users
        paginator = None
    else:
        paginator = Paginator(users, int(per_page))
        paginated_users = paginator.get_page(page_number)

        return render(request, 'admin/core/complete_user_data.html', {
        'users': paginated_users,
        'roles': roles,
        'provinces': provinces,
        'schooling_statuses': schooling_statuses,
        'paginator': paginator,
        'per_page': per_page,
        'today': date.today(),
    })

def bulk_upload_scq(request, bank_id):
    bank = QuestionBank.objects.get(id=bank_id)

    if bank.question_type != 'SCQ':
        messages.error(request, "Invalid bank type. This bank only supports SCQs.")
        return redirect('admin:core_questionbanktitle_changelist')

    if request.method == 'POST':
        form = UploadSCQForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            wb = openpyxl.load_workbook(file)
            sheet = wb.active

            rows_added = 0
            for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header
                question_text, option_1, option_2, option_3, option_4, correct_answer = row
                SCQQuestion.objects.create(
                    question_bank=bank,
                    question_text=question_text,
                    option_1=option_1,
                    option_2=option_2,
                    option_3=option_3,
                    option_4=option_4,
                    correct_answer=correct_answer
                )
                rows_added += 1

            messages.success(request, f"{rows_added} SCQ questions uploaded successfully!")
            return redirect(f'/preview-questions/{bank.id}/')
    else:
        form = UploadSCQForm(initial={'question_bank_id': bank.id})

    return render(request, 'admin/core/scq_upload_form.html', {'form': form, 'bank': bank})

def bulk_upload_mcq(request, bank_id):
    bank = QuestionBank.objects.get(id=bank_id)

    if bank.question_type != 'MCQ':
        messages.error(request, "Invalid bank type. This bank only supports MCQs.")
        return redirect('admin:core_questionbanktitle_changelist')

    if request.method == 'POST':
        form = UploadMCQForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            wb = openpyxl.load_workbook(file)
            sheet = wb.active

            rows_added = 0
            for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header
                question_text, option_1, option_2, option_3, option_4, correct_answers = row

                # Parse correct_answers like "A,C"
                correct_list = [opt.strip().upper() for opt in correct_answers.split(',') if opt.strip()]

                MCQQuestion.objects.create(
                    question_bank=bank,
                    question_text=question_text,
                    option_1=option_1,
                    option_2=option_2,
                    option_3=option_3,
                    option_4=option_4,
                    correct_answers=correct_list,
                )
                rows_added += 1

            messages.success(request, f"{rows_added} MCQ questions uploaded successfully!")
            return redirect(f'/preview-questions/{bank.id}/')
    else:
        form = UploadMCQForm(initial={'question_bank_id': bank.id})

    return render(request, 'admin/core/mcq_upload_form.html', {'form': form, 'bank': bank})

# core/admin_views.py

def bulk_upload_fib(request, bank_id):
    bank = QuestionBank.objects.get(id=bank_id)

    if bank.question_type != 'FIB':
        messages.error(request, "This bank only supports FIB questions.")
        return redirect('admin:core_questionbanktitle_changelist')

    if request.method == 'POST':
        form = UploadFIBForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            wb = openpyxl.load_workbook(file)
            sheet = wb.active

            rows_added = 0
            for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header
                question_text, correct_answers = row
                answer_list = [ans.strip() for ans in correct_answers.split(',') if ans.strip()]

                FIBQuestion.objects.create(
                    question_bank=bank,
                    question_text=question_text,
                    correct_answers=answer_list,
                )
                rows_added += 1

            messages.success(request, f"{rows_added} FIB questions uploaded successfully!")
            return redirect(f'/preview-questions/{bank.id}/')
    else:
        form = UploadFIBForm(initial={'question_bank_id': bank.id})

    return render(request, 'admin/core/fib_upload_form.html', {'form': form, 'bank': bank})


def preview_questions(request, bank_id):
    bank = get_object_or_404(QuestionBank, id=bank_id)
    questions = []

    if bank.type == 'SCQ':
        questions = SCQQuestion.objects.filter(question_bank=bank)
        question_type = 'SCQ'
    elif bank.type == 'MCQ':
        questions = MCQQuestion.objects.filter(question_bank=bank)
        question_type = 'MCQ'
    elif bank.type == 'FIB':
        questions = FIBQuestion.objects.filter(question_bank=bank)
        question_type = 'FIB'
    else:
        question_type = 'UNKNOWN'

    return render(request, 'admin/core/question_preview.html', {
        'bank': bank,
        'questions': questions,
        'question_type': question_type,
    })

@staff_member_required
def duplicate_question(request, question_type, question_id):
    model_map = {
        'SCQ': SCQQuestion,
        'MCQ': MCQQuestion,
        'FIB': FIBQuestion,
    }

    model = model_map.get(question_type.upper())
    if not model:
        messages.error(request, "Invalid question type.")
        return redirect('admin:index')

    original = get_object_or_404(model, id=question_id)
    original.pk = None  # Create a new object
    original.question_id = uuid.uuid4()
    original.save()

    messages.success(request, f"{question_type} Question duplicated successfully!")

    # Check if a safe redirect is requested
    redirect_to = request.GET.get('redirect_to')
    if redirect_to and redirect_to.startswith("/"):
        return redirect(redirect_to)

    # Default fallback if redirect_to not provided
    return redirect(f"/preview-questions/{original.question_bank.id}/")


@staff_member_required
def assign_questions_view(request, quiz_id):
    if request.user.role == 'manager':
        return HttpResponseForbidden("Managers are not allowed to assign questions to quizzes.")

    quiz = get_object_or_404(Quiz, id=quiz_id)
    question_banks = QuestionBank.objects.all().order_by(Lower('title'))  # Case-insensitive alphabetical sorting

    if request.method == 'POST':
        bank_id = int(request.POST.get("add_bank"))
        num_key = f"num_{bank_id}"
        try:
            num_questions = int(request.POST.get(num_key))
            bank = QuestionBank.objects.get(id=bank_id)

            # Check if assignment already exists
            existing = QuizQuestionAssignment.objects.filter(quiz=quiz, question_bank=bank).first()
            if existing:
                existing.num_questions = num_questions
                existing.save()
                messages.success(request, f"Updated existing assignment for '{bank.title}'")
            else:
                QuizQuestionAssignment.objects.create(
                    quiz=quiz,
                    question_bank=bank,
                    num_questions=num_questions
                )
                messages.success(request, f"Added '{bank.title}' to quiz.")

        except Exception as e:
            messages.error(request, f"Error: {str(e)}")

        return redirect(request.path)

    return render(request, "admin/core/assign_questions.html", {
        "quiz": quiz,
        "question_banks": question_banks
    })

@login_required
def admin_list_quizzes_view(request):
    if request.user.role != 'admin':
        return HttpResponseForbidden("Only admins can view this page.")

    # Read query parameters
    sort_field = request.GET.get('sort', 'title')
    sort_dir = request.GET.get('dir', 'asc')
    selected_grade = request.GET.get('grade')
    selected_subject = request.GET.get('subject')

    # Valid sortable fields
    valid_sort_fields = {
        'title': 'title',
        'grade': 'grade__name',
        'subject': 'subject__name',
        'chapter': 'chapter__name'
    }
    sort_expression = valid_sort_fields.get(sort_field, 'title')
    if sort_dir == 'desc':
        sort_expression = f'-{sort_expression}'

    # Base queryset
    quizzes = Quiz.objects.all().prefetch_related('assignments__question_bank', 'chapter', 'subject', 'grade')

    # Filtering
    if selected_grade:
        quizzes = quizzes.filter(grade__name=selected_grade)
    if selected_subject:
        quizzes = quizzes.filter(subject__name=selected_subject)

    # Sorting
    quizzes = quizzes.order_by(sort_expression)

    # Build data for display
    quiz_data = []
    for quiz in quizzes:
        assignments = quiz.assignments.all()
        question_banks = [a.question_bank.title for a in assignments]
        total_questions = sum(a.num_questions for a in assignments)

        quiz_data.append({
            'id': quiz.id,
            'title': quiz.title,
            'chapter': quiz.chapter.name if quiz.chapter else '',
            'subject': quiz.subject.name if quiz.subject else '',
            'grade': quiz.grade.name if quiz.grade else '',
            'marks_per_question': quiz.marks_per_question,
            'question_banks': question_banks,
            'total_questions': total_questions,
            'edit_url': f"/admin/core/quiz/{quiz.id}/change/",
            'delete_url': f"/admin/core/quiz/{quiz.id}/delete/",
            'format_url': reverse('quiz-question-assignments') + f"?quiz_id={quiz.id}",
            'attempt_url': f"/start-quiz/{quiz.id}/",
        })

    # Grade dropdown
    grades = Grade.objects.all().order_by('name')

    # Subject dropdown nested based on selected grade
    if selected_grade:
        grade_obj = Grade.objects.filter(name=selected_grade).first()
        subjects = Subject.objects.filter(grade=grade_obj).order_by('name') if grade_obj else Subject.objects.none()
    else:
        subjects = Subject.objects.all().order_by('name')

    # Table headers with (field, label) format for sortable columns
    headers = [
        ('title', 'Title'),
        ('grade', 'Grade'),
        ('subject', 'Subject'),
        ('chapter', 'Chapter'),
    ]

    return render(request, 'admin/core/list_quizzes.html', {
        'quiz_data': quiz_data,
        'grades': grades,
        'subjects': subjects,
        'request': request,
        'current_sort': sort_field,
        'current_dir': sort_dir,
        'headers': headers,  # ‚úÖ passed to template
    })

@staff_member_required
def list_backups(request):
    backup_dir = os.path.join(settings.MEDIA_ROOT, 'backups')
    os.makedirs(backup_dir, exist_ok=True)

    # ‚úÖ If user clicked "Create Backup" button
    if request.method == "POST":
        timestamp = timezone.localtime(timezone.now()).strftime("%Y-%m-%d_%H-%M-%S")
        filename = f"backup_{timestamp}.json"
        filepath = os.path.join(backup_dir, filename)

        with open(filepath, 'w') as f:
            call_command('dumpdata', '--natural-primary', '--natural-foreign', '--indent=2', stdout=f)

        # Keep only latest 3 backups
        files = sorted(
            [f for f in os.listdir(backup_dir) if f.endswith('.json')],
            reverse=True
        )
        for old_file in files[3:]:
            os.remove(os.path.join(backup_dir, old_file))

        return HttpResponseRedirect(request.path)  # Refresh page after backup

    # ‚úÖ Show existing backups
    backup_files = [
        f for f in os.listdir(backup_dir) if f.endswith('.json')
    ]
    backup_files.sort(reverse=True)

    return render(request, 'admin/backups.html', {
        'backup_files': backup_files
    })


@staff_member_required
def download_backup(request, filename):
    backup_dir = os.path.join(settings.MEDIA_ROOT, 'backups')
    filepath = os.path.join(backup_dir, filename)
    return FileResponse(open(filepath, 'rb'), as_attachment=True)

@user_passes_test(lambda u: u.is_superuser)
@require_POST
def restore_backup(request, filename):
    backup_dir = os.path.join(settings.MEDIA_ROOT, 'backups')  # ‚úÖ FIXED
    filepath = os.path.join(backup_dir, filename)

    if not os.path.exists(filepath):
        messages.error(request, "Backup file not found.")
    else:
        try:
            call_command('loaddata', filepath)
            messages.success(request, f"{filename} restored successfully.")
        except Exception as e:
            messages.error(request, f"Restore failed: {str(e)}")

    return redirect('list_backups')

@staff_member_required
def user_dashboard(request):
    return render(request, 'admin/dashboard/admin_users.html')

@staff_member_required
def admin_question_bank_view(request):
    # Fetch all question banks
    question_banks = QuestionBank.objects.all().order_by('-id')

    # Create a dictionary to hold question counts by bank_id
    question_counts = {}

    from core.models import SCQQuestion, MCQQuestion, FIBQuestion  # Adjust if models live elsewhere

    # Count SCQ questions
    for bank_id, count in SCQQuestion.objects.values_list('question_bank',).annotate(c=Count('id')):
        question_counts[bank_id] = question_counts.get(bank_id, 0) + count

    # Count MCQ questions
    for bank_id, count in MCQQuestion.objects.values_list('question_bank',).annotate(c=Count('id')):
        question_counts[bank_id] = question_counts.get(bank_id, 0) + count

    # Count FIB questions
    for bank_id, count in FIBQuestion.objects.values_list('question_bank',).annotate(c=Count('id')):
        question_counts[bank_id] = question_counts.get(bank_id, 0) + count

    # Attach count to each question bank object
    for bank in question_banks:
        bank.question_count = question_counts.get(bank.id, 0)

    return render(request, "admin/dashboard/admin_question_bank.html", {
        "question_banks": question_banks
    })

@staff_member_required
def delete_question(request, q_type, q_id, bank_id):
    if q_type == 'SCQ':
        SCQQuestion.objects.filter(question_id=q_id).delete()
    elif q_type == 'MCQ':
        MCQQuestion.objects.filter(question_id=q_id).delete()
    elif q_type == 'FIB':
        FIBQuestion.objects.filter(question_id=q_id).delete()

    return redirect(f'/preview-questions/{bank_id}/')


@staff_member_required
def edit_question_bank(request, bank_id):
    bank = get_object_or_404(QuestionBank, id=bank_id)

    if request.method == 'POST':
        bank.title = request.POST.get('title')
        bank.type = request.POST.get('type')
        bank.save()
        messages.success(request, 'Question Bank updated successfully.')
        return redirect('admin-question-bank')

    return render(request, 'admin/core/edit_question_bank.html', {'bank': bank})

@staff_member_required
def delete_question_bank(request, bank_id):
    bank = get_object_or_404(QuestionBank, id=bank_id)
    if request.method == 'POST':
        bank.delete()
        messages.success(request, 'Question Bank deleted successfully.')
        return redirect('admin-question-bank')

    return render(request, 'admin/core/confirm_delete_question_bank.html', {'bank': bank})

@staff_member_required
def admin_quiz_dashboard(request):
    return render(request, 'admin/dashboard/admin_quizzes.html')

@staff_member_required
def create_metadata_view(request):
    grades = Grade.objects.all().order_by('name')
    subjects = Subject.objects.all().order_by('grade__name', 'name')
    chapters = Chapter.objects.all().order_by('subject__name', 'name')

    return render(request, 'admin/core/create_metadata.html', {
        'grades': grades,
        'subjects': subjects,
        'chapters': chapters
    })

@staff_member_required
def quiz_question_assignment_view(request):
    quizzes = Quiz.objects.all()
    grades = Grade.objects.all()
    
    # √î¬£√∏‚àö¬∫‚àö¬®‚àö‚Ä¢ Apply grade filter if selected
    selected_grade_id = request.GET.get('grade_filter')
    if selected_grade_id:
        quizzes = quizzes.filter(grade_id=selected_grade_id)

    question_banks = QuestionBank.objects.all().order_by(Lower('title'))  # ‚Äö√Ñ√∂‚àö‚à´‚àö√± Alphabetical sorting

    if request.method == 'POST':
        quiz_id = request.POST.get("quiz_id")
        bank_id = request.POST.get("assign_bank_id")
        num_key = f"num_questions_{bank_id}"
        num_questions = request.POST.get(num_key)

        if quiz_id and bank_id and num_questions:
            try:
                quiz = Quiz.objects.get(id=quiz_id)
                bank = QuestionBank.objects.get(id=bank_id)
                num_questions = int(num_questions)

                QuizQuestionAssignment.objects.create(
                    quiz=quiz,
                    question_bank=bank,
                    num_questions=num_questions
                )
                messages.success(request, f" {num_questions} questions from '{bank.title}' assigned to quiz '{quiz.title}'.")
                return redirect('quiz-question-assignments')  # ‚Äö√Ñ√∂‚àö‚à´‚àö√± works now
            except Exception as e:
                messages.error(request, f"‚Äö√Ñ√∂‚àöœÄ‚àö‚Ä¢ Error: {e}")

    return render(request, 'admin/core/quiz_question_assignments.html', {
        'quizzes': quizzes,
        'grades': grades,
        'selected_grade_id': selected_grade_id,
        'question_banks': question_banks
    })


# KEEP this in admin_views.py and update it to include line_spacing
class QuizFormattingForm(forms.ModelForm):
    class Meta:
        model = Quiz
        fields = ['input_box_width', 'text_alignment', 'font_size', 'line_spacing']
        widgets = {
            'input_box_width': forms.NumberInput(attrs={'min': 1}),
            'text_alignment': forms.Select(choices=[('left', 'Left'), ('center', 'Center'), ('right', 'Right')]),
            'font_size': forms.NumberInput(attrs={'min': 10, 'max': 40}),
            'line_spacing': forms.NumberInput(attrs={'step': 0.1, 'min': 1, 'max': 3}),
        }

@staff_member_required
def quiz_formatting_view(request, quiz_id):
    quiz = get_object_or_404(Quiz, id=quiz_id)
    if request.method == 'POST':
        form = QuizFormattingForm(request.POST, instance=quiz)
        if form.is_valid():
            form.save()
            return redirect('admin-list-quizzes')
    else:
        form = QuizFormattingForm(instance=quiz)

    return render(request, 'admin/core/quiz_formatting.html', {
        'form': form,
        'quiz': quiz
    })


@login_required
@user_passes_test(lambda u: u.is_authenticated and u.role == 'admin')
def stats_dashboard_view(request):
    context = {
        'descriptive_stats': {},
        'national_overview': {},
        'provincial_overview': [],
        'gender_language_crosstab': [],
    }
    return render(request, 'admin/core/stats_dashboard.html', context)

User = get_user_model()

@require_POST
@staff_member_required
def bulk_delete_users(request):
    selected_ids = request.POST.getlist('selected_users')

    if not selected_ids:
        messages.warning(request, "‚ö†Ô∏è No users were selected for deletion.")
        return redirect(request.META.get('HTTP_REFERER', '/admin/core/user/complete_user_data/'))

    # Exclude admin users from deletion
    users_to_delete = User.objects.filter(id__in=selected_ids).exclude(role='admin')
    deleted_count = users_to_delete.count()

    # Perform deletion
    users_to_delete.delete()

    # Success message
    messages.success(request, f"‚úÖ {deleted_count} user(s) deleted successfully.")

    # Redirect back to the same filtered view
    return redirect(request.META.get('HTTP_REFERER', '/admin/core/user/complete_user_data/'))

def get_subjects_by_grade(request):
    grade_id = request.GET.get('grade_id')
    if grade_id:
        subjects = Subject.objects.filter(grade_id=grade_id).values('id', 'name')
        return JsonResponse(list(subjects), safe=False)
    return JsonResponse([], safe=False)

def get_chapters_by_subject(request):
    subject_id = request.GET.get('subject_id')
    if subject_id:
        chapters = Chapter.objects.filter(subject_id=subject_id).values('id', 'name')
        return JsonResponse(list(chapters), safe=False)
    return JsonResponse([], safe=False)
